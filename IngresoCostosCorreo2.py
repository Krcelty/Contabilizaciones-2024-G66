import imaplib
import email
from email.header import decode_header
import openpyxl
import re
import pandas as pd

# Parámetros de conexión
IMAP_SERVER = "imap.gmail.com"  # Cambiar al servidor IMAP correcto
EMAIL_ACCOUNT = "constanza.perez@global66.com"
PASSWORD = "oxen tifh gizb sgtm"
EMAIL_SUBJECT = "Ingresos y costos operativos - Octubre 2024"  

# Crear un DataFrame vacío con la estructura especificada
columns = [
    'Tipo de comprobante', 'Esquema', 'Glosa comprobante', 'Fecha contable', 
    'Total debe', 'Total haber', 'Ítem detalle comprobante', 'Código unidad de negocio', 
    'Glosa comprobante', 'RUT cliente', 'RUT personal', 'Código cuenta contable', 
    'Código centro costos', 'Monto debe', 'Monto haber', 'Tipo de documento', 
    'Fecha de documento', 'Número de documento', 'Concepto 1', 'Concepto 2', 
    'Concepto 3', 'Concepto 4', 'Tipo contabilidad'
]

output_df = pd.DataFrame(columns=columns)

def conectar_email():
    # Conectar a la bandeja de entrada de IMAP
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ACCOUNT, PASSWORD)
    return mail

def buscar_correo(mail, asunto):
    # Seleccionar la bandeja de entrada
    mail.select("inbox")
    # Buscar correo con el asunto específico
    status, mensajes = mail.search(None, f'SUBJECT "{asunto}"')
    if status != "OK":
        print("No se encontraron correos con el asunto especificado.")
        return None

    # Obtener el ID del primer correo que coincide con el asunto
    mensaje_ids = mensajes[0].split()
    if mensaje_ids:
        return mensaje_ids[0]
    else:
        return None

def leer_correo(mail, mensaje_id):
    # Obtener el correo completo por su ID
    status, mensaje_datos = mail.fetch(mensaje_id, "(RFC822)")
    if status != "OK":
        print("Error al obtener el correo.")
        return None

    # Decodificar y analizar el mensaje
    mensaje = email.message_from_bytes(mensaje_datos[0][1])
    contenido = ""

    # Si el correo tiene varias partes (texto y HTML, por ejemplo)
    if mensaje.is_multipart():
        for parte in mensaje.walk():
            # Usar solo la parte de texto (descartar HTML)
            if parte.get_content_type() == "text/plain":
                contenido += parte.get_payload(decode=True).decode('utf-8', errors='ignore')
    else:
        contenido = mensaje.get_payload(decode=True).decode('utf-8', errors='ignore')

    return contenido

def extraer_datos(texto):
    # Expresión regular para buscar el monto y el código de cuenta contable
    regex = r'(\(Cuenta (\d+)\):\s([\d\.,]+)\sCLP)'
    datos = []
    
    matches = re.findall(regex, texto)
    for match in matches:
        cuenta_contable = match[1]
        monto = float(match[2].replace('.', '').replace(',', '.'))  # Convertir monto a float
        datos.append((cuenta_contable, monto))
    
    return datos

def guardar_en_excel(df):
    # Crear un archivo Excel y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contenido del Correo"

    # Escribir los encabezados
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Escribir los datos del DataFrame
    for row_num, row_data in df.iterrows():
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num + 2, column=col_num, value=cell_value)
    
    # Guardar el archivo
    wb.save("contenido_correo.xlsx")
    print("Archivo Excel creado con éxito.")

def main():
    # Conectar al servidor de correo y buscar el correo deseado
    mail = conectar_email()
    mensaje_id = buscar_correo(mail, EMAIL_SUBJECT)

    if mensaje_id:
        # Leer el correo
        contenido = leer_correo(mail, mensaje_id)
        
        # Extraer datos (cuenta contable y monto)
        if contenido:
            datos = extraer_datos(contenido)

            # Crear DataFrame y llenar con los datos extraídos
            for cuenta_contable, monto in datos:
                output_df = output_df.append({
                    'Tipo de comprobante': 'T',  # Puedes ajustar según corresponda
                    'Esquema': 'Esquema',  # Ejemplo
                    'Glosa comprobante': 'Ingresos remesas',  # Ejemplo
                    'Fecha contable': '2024-10-01',  # Ajusta la fecha según corresponda
                    'Total debe': 0,
                    'Total haber': monto,
                    'Ítem detalle comprobante': 1,  # Incrementa según necesites
                    'Código unidad de negocio': 1,
                    'Glosa comprobante': 'Ingresos remesas',
                    'RUT cliente': '',
                    'RUT personal': '',
                    'Código cuenta contable': cuenta_contable,
                    'Código centro costos': '',
                    'Monto debe': 0,
                    'Monto haber': monto,
                    'Fecha de documento': '2024-10-01',
                    'Número de documento': '',
                    'Concepto 1': '',
                    'Concepto 2': '',
                    'Concepto 3': '',
                    'Concepto 4': '',
                    'Tipo contabilidad': ''
                }, ignore_index=True)

            # Guardar el DataFrame en un archivo Excel
            guardar_en_excel(output_df)
    else:
        print("No se encontró el correo con el asunto especificado.")

    # Cerrar la conexión
    mail.logout()

if __name__ == "__main__":
    main()
